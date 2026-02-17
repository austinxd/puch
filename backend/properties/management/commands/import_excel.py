from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from properties.models import Agent, Property


# Default column mapping (Excel column header -> model field)
DEFAULT_MAPPING = {
    'identificador': 'identificador',
    'clase': 'clase',
    'agente': '_agent_name',
    'nombre': 'nombre',
    'tipologia': 'tipologia',
    'operacion': 'operacion',
    'link maps': 'link_maps',
    'distrito': 'distrito',
    'pitch': 'pitch',
    'calle': 'calle',
    'direccion': 'direccion',
    'referencia': 'referencia',
    'antiguedad': 'antiguedad',
    'precio': 'precio',
    'costo mantenimiento': 'costo_mantenimiento',
    'metraje': 'metraje',
    'vista': 'vista',
    'distribucion': 'distribucion',
    'ascensor': 'ascensor',
    'habitaciones': 'habitaciones',
    'cocheras': 'cocheras',
    'cantidad pisos': 'cantidad_pisos',
    'tipo cocina': 'tipo_cocina',
    'terraza balcon': 'terraza_balcon',
    'piso': 'piso',
    'baños': 'banos',
    'cuarto servicio': 'cuarto_servicio',
    'baño servicio': 'bano_servicio',
    'documentacion': 'documentacion',
    'parametros usos': 'parametros_usos',
    'financiamiento': 'financiamiento',
    'imagen 1': 'imagen_1',
    'imagen 2': 'imagen_2',
    'imagen 3': 'imagen_3',
    'imagen 4': 'imagen_4',
    'imagen 5': 'imagen_5',
    'video': 'video',
    'recorrido 360': 'recorrido_360',
}


class Command(BaseCommand):
    help = 'Import properties from an Excel file'

    def add_arguments(self, parser):
        parser.add_argument('file', type=str, help='Path to the Excel file')
        parser.add_argument(
            '--sheet', type=str, default=None,
            help='Sheet name (defaults to first sheet)',
        )
        parser.add_argument(
            '--dry-run', action='store_true',
            help='Preview without saving',
        )

    def handle(self, *args, **options):
        filepath = options['file']
        sheet_name = options['sheet']
        dry_run = options['dry_run']

        wb = load_workbook(filepath, read_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        self.stdout.write(f'Reading sheet: {ws.title}')

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            self.stdout.write(self.style.WARNING('Empty sheet'))
            return

        # Build header mapping
        raw_headers = [str(h).strip().lower() if h else '' for h in rows[0]]
        col_map = {}
        for idx, header in enumerate(raw_headers):
            if header in DEFAULT_MAPPING:
                col_map[idx] = DEFAULT_MAPPING[header]
            else:
                # Try partial match
                for key, field in DEFAULT_MAPPING.items():
                    if key in header or header in key:
                        col_map[idx] = field
                        break

        if not col_map:
            self.stdout.write(self.style.ERROR('Could not map any columns. Headers found:'))
            self.stdout.write(str(raw_headers))
            return

        self.stdout.write(f'Mapped columns: {len(col_map)}')

        created = 0
        updated = 0
        errors = 0

        for row_num, row in enumerate(rows[1:], start=2):
            data = {}
            agent_name = None

            for col_idx, field_name in col_map.items():
                value = row[col_idx] if col_idx < len(row) else None
                if value is None:
                    continue
                value = str(value).strip() if not isinstance(value, (int, float)) else value

                if field_name == '_agent_name':
                    agent_name = str(value).strip() if value else None
                    continue

                if field_name == 'precio':
                    try:
                        data[field_name] = float(str(value).replace(',', '').replace('$', '').replace('S/', '').strip())
                    except (ValueError, TypeError):
                        data[field_name] = None
                    continue

                data[field_name] = str(value) if value else ''

            identificador = data.get('identificador', '').strip()
            if not identificador:
                continue

            if dry_run:
                self.stdout.write(f'  [DRY RUN] Row {row_num}: {identificador} - {data.get("nombre", "")}')
                created += 1
                continue

            # Get or create agent
            agent = None
            if agent_name:
                agent, _ = Agent.objects.get_or_create(name=agent_name)

            data['agent'] = agent

            try:
                obj, was_created = Property.objects.update_or_create(
                    identificador=identificador,
                    defaults=data,
                )
                if was_created:
                    created += 1
                else:
                    updated += 1
            except Exception as e:
                errors += 1
                self.stdout.write(self.style.ERROR(f'  Row {row_num}: {e}'))

        wb.close()
        action = '[DRY RUN] Would create' if dry_run else 'Created'
        self.stdout.write(self.style.SUCCESS(
            f'\nDone! {action}: {created}, Updated: {updated}, Errors: {errors}'
        ))
