"""
Script to upload properties from the Brikia Excel file to the API.
Usage: python scripts/upload_properties.py <excel_file> <api_base_url>
Example: python scripts/upload_properties.py "/Users/austin/Downloads/IA Brikia .xlsx" https://api.brikia.tech/api
"""
import re
import sys
import requests
from openpyxl import load_workbook


def extract_url(value):
    """Extract URL from markdown image format ![name](url) or return as-is."""
    if not value:
        return ''
    value = str(value).strip()
    match = re.search(r'\[.*?\]\((.*?)\)', value)
    if match:
        return match.group(1)
    return value


def parse_phone(value):
    """Clean phone number from Excel formula like =51961730160."""
    if not value:
        return ''
    value = str(value).strip()
    value = value.lstrip('=')
    if value.isdigit():
        return f'+{value}'
    return value


def main():
    if len(sys.argv) < 3:
        print("Usage: python upload_properties.py <excel_file> <api_base_url>")
        print('Example: python upload_properties.py "file.xlsx" https://api.brikia.tech/api')
        sys.exit(1)

    filepath = sys.argv[1]
    base_url = sys.argv[2].rstrip('/')

    wb = load_workbook(filepath, read_only=True)
    ws = wb['Sheet1']
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]

    print(f"Found {len(rows) - 1} rows (including empty)")

    # Column indices based on the Excel structure
    COL = {
        'identificador': 0,
        'clase': 1,
        'owner': 2,
        'telf': 3,
        'correo': 4,
        'nombre': 5,
        'tipologia': 6,
        'operacion': 7,
        'link_maps': 8,
        'distrito': 9,
        'pitch': 10,
        'calle': 11,
        'direccion': 12,
        'referencia': 13,
        'antiguedad': 14,
        'precio': 15,
        'costo_mantenimiento': 16,
        'metraje': 17,
        'vista': 18,
        'distribucion': 19,
        'ascensor': 20,
        'habitaciones': 21,
        'cocheras': 22,
        'cantidad_pisos': 23,
        'tipo_cocina': 24,
        'terraza_balcon': 25,
        'piso': 26,
        'banos': 27,
        'cuarto_servicio': 28,
        'bano_servicio': 29,
        'documentacion': 30,
        'parametros_usos': 31,
        'financiamiento': 32,
        'imagen_1': 33,
        'imagen_2': 34,
        'imagen_3': 35,
        'imagen_4': 36,
        'imagen_5': 37,
        'video': 38,
        'recorrido_360': 39,
    }

    def get(row, key):
        idx = COL[key]
        if idx >= len(row):
            return None
        return row[idx]

    def s(val):
        """Safely convert to string."""
        if val is None:
            return ''
        return str(val).strip()

    # Cache for agents: name -> id
    agents_cache = {}

    created = 0
    errors = 0

    for row_num, row in enumerate(rows[1:], start=2):
        if not row or not get(row, 'identificador'):
            continue

        identificador = s(get(row, 'identificador'))
        print(f"\n--- Row {row_num}: {identificador} ---")

        # Create or get agent
        owner = s(get(row, 'owner'))
        agent_id = None
        if owner:
            if owner not in agents_cache:
                # Check if agent exists
                res = requests.get(f"{base_url}/agents/", params={"search": owner})
                existing = res.json().get('results', [])
                match = next((a for a in existing if a['name'] == owner), None)
                if match:
                    agents_cache[owner] = match['id']
                    print(f"  Agent found: {owner} (id={match['id']})")
                else:
                    # Create agent
                    agent_data = {
                        'name': owner,
                        'phone': parse_phone(get(row, 'telf')),
                        'email': s(get(row, 'correo')),
                    }
                    res = requests.post(f"{base_url}/agents/", json=agent_data)
                    if res.status_code == 201:
                        agents_cache[owner] = res.json()['id']
                        print(f"  Agent created: {owner} (id={res.json()['id']})")
                    else:
                        print(f"  ERROR creating agent: {res.text}")
            agent_id = agents_cache.get(owner)

        # Parse price
        precio_raw = get(row, 'precio')
        precio = None
        if precio_raw is not None:
            try:
                precio = float(str(precio_raw).replace(',', '').replace('$', '').replace('S/', '').strip())
            except (ValueError, TypeError):
                pass

        # Parse costo_mantenimiento
        costo = get(row, 'costo_mantenimiento')
        if costo is not None and isinstance(costo, (int, float)):
            costo = str(int(costo)) if costo == int(costo) else str(costo)
        else:
            costo = s(costo)

        property_data = {
            'identificador': identificador,
            'clase': s(get(row, 'clase')),
            'agent': agent_id,
            'nombre': s(get(row, 'nombre')),
            'tipologia': s(get(row, 'tipologia')),
            'operacion': s(get(row, 'operacion')),
            'link_maps': s(get(row, 'link_maps')),
            'distrito': s(get(row, 'distrito')),
            'pitch': s(get(row, 'pitch')),
            'calle': s(get(row, 'calle')),
            'direccion': s(get(row, 'direccion')),
            'referencia': s(get(row, 'referencia')),
            'antiguedad': s(get(row, 'antiguedad')),
            'precio': precio,
            'costo_mantenimiento': costo,
            'metraje': s(get(row, 'metraje')),
            'vista': s(get(row, 'vista')),
            'distribucion': s(get(row, 'distribucion')),
            'ascensor': s(get(row, 'ascensor')),
            'habitaciones': s(get(row, 'habitaciones')),
            'cocheras': s(get(row, 'cocheras')),
            'cantidad_pisos': s(get(row, 'cantidad_pisos')),
            'tipo_cocina': s(get(row, 'tipo_cocina')),
            'terraza_balcon': s(get(row, 'terraza_balcon')),
            'piso': s(get(row, 'piso')),
            'banos': s(get(row, 'banos')),
            'cuarto_servicio': s(get(row, 'cuarto_servicio')),
            'bano_servicio': s(get(row, 'bano_servicio')),
            'documentacion': s(get(row, 'documentacion')),
            'parametros_usos': s(get(row, 'parametros_usos')),
            'financiamiento': s(get(row, 'financiamiento')),
            'imagen_1': extract_url(get(row, 'imagen_1')),
            'imagen_2': extract_url(get(row, 'imagen_2')),
            'imagen_3': extract_url(get(row, 'imagen_3')),
            'imagen_4': extract_url(get(row, 'imagen_4')),
            'imagen_5': extract_url(get(row, 'imagen_5')),
            'video': extract_url(get(row, 'video')),
            'recorrido_360': s(get(row, 'recorrido_360')),
            'activo': True,
        }

        res = requests.post(f"{base_url}/properties/", json=property_data)
        if res.status_code == 201:
            print(f"  Created: {property_data['nombre']}")
            created += 1
        else:
            print(f"  ERROR: {res.status_code} - {res.text[:300]}")
            errors += 1

    wb.close()
    print(f"\n{'='*40}")
    print(f"Done! Created: {created}, Errors: {errors}")


if __name__ == '__main__':
    main()
